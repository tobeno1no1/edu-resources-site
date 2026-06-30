#!/usr/bin/env python3
"""
从 Excel 文件智能解析标题并批量生成教辅资料 markdown 文件
用法：python batch_import_excel.py <excel_path>
"""

import os
import sys
import re
from datetime import datetime, timezone, timedelta

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl，请安装：pip install openpyxl")
    sys.exit(1)

CONTENT_DIR = os.path.join(os.path.dirname(__file__), '..', 'src', 'content', 'resources')

# ─── 年级解析 ────────────────────────────────────────────
GRADE_PATTERNS = [
    # "X升Y" → Y（暑假衔接用的是升入年级）
    (re.compile(r'([一二三四五六])升([七八九十])'), lambda m: _ch2num(m.group(2))),
    (re.compile(r'([一二三四五六七八九])升([一二三四五六七八九])'), lambda m: _ch2num(m.group(2))),
    # "幼升小" → 1
    (re.compile(r'幼升小'), lambda m: 1),
    # "小升初" → 7
    (re.compile(r'小升初'), lambda m: 7),
    # "X年级"
    (re.compile(r'([一二三四五六七八九])年级'), lambda m: _ch2num(m.group(1))),
    # 数字年级
    (re.compile(r'(\d)年级'), lambda m: int(m.group(1))),
    # "初一/初二/初三"
    (re.compile(r'初一'), lambda m: 7),
    (re.compile(r'初二'), lambda m: 8),
    (re.compile(r'初三'), lambda m: 9),
    # "七年级/八年级/九年级"
    (re.compile(r'(七|八|九)年级'), lambda m: {'七':7,'八':8,'九':9}[m.group(1)]),
    # "1-6年级" 范围 → 返回列表
    (re.compile(r'(\d)[\-~到](\d)年级'), lambda m: list(range(int(m.group(1)), int(m.group(2))+1))),
    (re.compile(r'([一二三四五六])[-~到]([六七八九十])年级'), lambda m: list(range(_ch2num(m.group(1)), _ch2num(m.group(2))+1))),
]

CH_NUM = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10}

def _ch2num(s):
    if s in CH_NUM: return CH_NUM[s]
    try: return int(s)
    except: return 0

def parse_grade(title):
    """从标题中解析年级，返回 list[int]（支持范围拆分）"""
    for pat, handler in GRADE_PATTERNS:
        m = pat.search(title)
        if m:
            result = handler(m)
            if isinstance(result, list):
                return result
            return [result] if result else []
    return [0]  # 无明确年级

# ─── 科目解析 ────────────────────────────────────────────
SUBJECT_MAP = {
    '语文': '语文', '数学': '数学', '英语': '英语',
    '物理': '物理', '化学': '化学', '生物': '生物',
    '历史': '历史', '地理': '地理',
    '道德与法治': '道德与法治', '道法': '道德与法治',
    '政治': '道德与法治',
    '科学': '科学',
}

def parse_subject(title):
    # 先匹配长词（道德与法治 > 道法 > 政治）
    for key in sorted(SUBJECT_MAP.keys(), key=len, reverse=True):
        if key in title:
            return SUBJECT_MAP[key]
    return '综合'

# ─── 品牌/版本解析 ──────────────────────────────────────
BRAND_MAP = {
    '人教PEP版': '人教PEP版',
    '人教版': '人教版',
    '部编版': '部编版',
    '苏教版': '苏教版',
    '北师大版': '北师大版',
    '统编版': '部编版',
}

def parse_brand(title, subject):
    for key in sorted(BRAND_MAP.keys(), key=len, reverse=True):
        if key in title:
            return BRAND_MAP[key]
    # 默认推断
    if subject == '语文': return '部编版'
    if subject == '数学': return '人教版'
    if subject == '英语': return '人教PEP版'
    if subject == '科学': return '苏教版'
    return '通用'

# ─── 类型解析 ────────────────────────────────────────────
TYPE_RULES = [
    (re.compile(r'测试卷|期末测试'), '测试卷'),
    (re.compile(r'真题卷|押题|押题预测'), '真题'),
    (re.compile(r'暑假作业'), '暑假作业'),
    (re.compile(r'暑假预习|预习'), '预习资料'),
    (re.compile(r'暑假衔接|衔接'), '暑假衔接'),
    (re.compile(r'课课贴'), '课课贴'),
    (re.compile(r'字贴|写字表|练字|练字打卡'), '字贴练字'),
    (re.compile(r'口算题卡|口算|计算专项|脱式计算|竖式计算|简便计算|每日一练'), '专项练习'),
    (re.compile(r'知识点汇总|必背内容|高频考点|考点汇总|考点归纳|核心考点|知识点清单|重要知识点'), '知识点汇总'),
    (re.compile(r'专项练习|专项训练|专项|易错题|选择题'), '专项练习'),
    (re.compile(r'阅读理解|课外阅读'), '专项练习'),
    (re.compile(r'作文|范文|写作'), '专项练习'),
    (re.compile(r'练习|训练|冲刺'), '专项练习'),
]

def parse_type(title):
    for pat, label in TYPE_RULES:
        if pat.search(title):
            return label
    return '综合资料'

# ─── 标题清理 ────────────────────────────────────────────
def clean_title(title):
    """去掉常见后缀噪音"""
    title = title.strip()
    # 去掉末尾的"  电子版可打印"、"pdf电子版可打印"、"电子版可下载可打印"等
    title = re.sub(r'\s*(pdf|PDF)?电子版可[下载打印]+(\！)?$', '', title)
    title = re.sub(r'\s*完整电子版可打印$', '', title)
    title = re.sub(r'\s*高清电子版可打印$', '', title)
    title = re.sub(r'\s*/\s*读测试题', '', title)
    # 去掉多余空格
    title = re.sub(r'\s{2,}', ' ', title)
    return title.strip()

# ─── 日期 ─────────────────────────────────────────────────
def get_today():
    tz = timezone(timedelta(hours=8))
    return datetime.now(tz).strftime('%Y-%m-%d')

# ─── 生成 filename ────────────────────────────────────────
SUBJECT_ABBR = {
    '语文': 'cn', '数学': 'math', '英语': 'en',
    '物理': 'phy', '化学': 'chem', '生物': 'bio',
    '历史': 'hist', '地理': 'geo', '道德与法治': 'daofa',
    '科学': 'sci', '综合': 'misc',
}

def slugify(text):
    return re.sub(r'[^\w\-]', '', text.replace(' ', '-'))[:30]

counter = {}

def generate_filename(date, subject, grade, title):
    subj = SUBJECT_ABBR.get(subject, slugify(subject))
    key = f"{date}-{subj}-g{grade}"
    counter[key] = counter.get(key, 0) + 1
    return f"{date}-{subj}-g{grade}-{counter[key]:03d}.md"

# ─── 生成 markdown ────────────────────────────────────────
def generate_md(title, date, grade, subject, brand, rtype, link, tags, description):
    tags_yaml = ', '.join(f'"{t.strip()}"' for t in tags if t.strip())
    desc = description.replace('"', '\\"')
    return f"""---
title: "{title}"
date: "{date}"
grade: "{grade}"
subject: "{subject}"
brand: "{brand}"
type: "{rtype}"
link: "{link}"
tags: [{tags_yaml}]
description: "{desc}"
---
"""

# ─── 主流程 ────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("用法: python batch_import_excel.py <excel_path>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    date = get_today()
    added = []
    skipped = []

    for row_idx in range(2, ws.max_row + 1):
        raw_title = ws.cell(row_idx, 2).value
        link = ws.cell(row_idx, 8).value

        if not raw_title or not link:
            skipped.append((row_idx, '标题或链接为空'))
            continue

        raw_title = str(raw_title).strip()
        link = str(link).strip()
        title = clean_title(raw_title)

        # 解析字段
        grades = parse_grade(title)
        subject = parse_subject(title)
        brand = parse_brand(title, subject)
        rtype = parse_type(title)

        # 高中资料 → grade=10
        if '高中' in title:
            grades = [10]

        # 年级范围资料 → 拆成多条
        for grade in grades:
            if grade == 0:
                # 无明确年级的资料（如暑假练字），用 grade=0
                pass
            if grade > 9 and grade != 10:
                skipped.append((row_idx, f'年级超出范围: {grade}'))
                continue

            # 对范围资料，标题里替换年级范围为具体年级
            if len(grades) > 1:
                # 如果标题里是"1-6年级"，替换为具体年级
                grade_title = title
                range_pattern = re.compile(r'(\d)[\-~到](\d)年级')
                m = range_pattern.search(grade_title)
                if m:
                    grade_ch = {1:'一',2:'二',3:'三',4:'四',5:'五',6:'六',7:'七',8:'八',9:'九'}
                    grade_title = grade_title.replace(m.group(0), f'{grade_ch.get(grade, str(grade))}年级')
                # 中文范围
                ch_range_pattern = re.compile(r'([一二三四五六])[-~到]([六七八九十])年级')
                m2 = ch_range_pattern.search(grade_title)
                if m2:
                    grade_ch = {1:'一',2:'二',3:'三',4:'四',5:'五',6:'六',7:'七',8:'八',9:'九'}
                    grade_title = grade_title.replace(m2.group(0), f'{grade_ch.get(grade, str(grade))}年级')
            else:
                grade_title = title

            filename = generate_filename(date, subject, grade, grade_title)
            filepath_out = os.path.join(CONTENT_DIR, filename)

            # 检查是否已存在（忽略覆盖）
            if os.path.exists(filepath_out):
                skipped.append((row_idx, f'文件已存在: {filename}'))
                continue

            tags = []
            # 自动添加标签
            if '暑假' in grade_title:
                tags.append('暑假')
            if '期末' in grade_title:
                tags.append('期末复习')
            if '下册' in grade_title:
                tags.append('下册')
            elif '上册' in grade_title:
                tags.append('上册')

            content = generate_md(
                grade_title, date, str(grade), subject, brand, rtype, link,
                tags, ''
            )

            os.makedirs(CONTENT_DIR, exist_ok=True)
            with open(filepath_out, 'w', encoding='utf-8') as f:
                f.write(content)

            added.append((filename, grade_title, grade, subject))

    # 输出统计
    print(f"\n{'='*50}")
    print(f"✅ 成功添加 {len(added)} 份资料")
    print(f"⚠️  跳过 {len(skipped)} 条")
    if skipped:
        for idx, reason in skipped[:5]:
            print(f"   行{idx}: {reason}")
        if len(skipped) > 5:
            print(f"   ... 还有 {len(skipped)-5} 条跳过")
    print(f"{'='*50}")

    # 按年级统计
    grade_stats = {}
    for _, _, grade, subject in added:
        grade_label = f"年级{grade}" if grade else "跨年级"
        key = f"{grade_label}-{subject}"
        grade_stats[key] = grade_stats.get(key, 0) + 1
    print("\n📊 添加统计：")
    for key, count in sorted(grade_stats.items()):
        print(f"   {key}: {count}条")

    print(f"\n📌 推送到 GitHub 更新网站：")
    print("   git add src/content/resources/")
    print(f'   git commit -m "批量新增 {len(added)} 份资料 ({date})"')
    print("   git push origin main")


if __name__ == '__main__':
    main()
